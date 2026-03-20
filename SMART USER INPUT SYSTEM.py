import json
import os.path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
import datetime

def is_valid_email(email):
    pattern=r'^[\w\.-]+@[\w\.-]+\.\w+$'
    return re.match(pattern, email)
def delete_user(name, age, email):
    global USER
    for user in USER:
        if user["name"] == name and user["age"] == age and user["email"] == email:
            USER.remove(user)
            print("User deleted successfully")
            return
    print("User not found")
def save_users_to_excel(USER,  filename = f"USER_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"):
    wb = openpyxl.Workbook()
    ws=wb.active
    ws.title="USER"
    headers=["Name","Age","Email"]
    #header
    for col,header in enumerate(headers,start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill= PatternFill(fill_type="solid", start_color="4472C4")
        cell.alignment = Alignment(horizontal="center")
    #cell data
    for row, users in enumerate(USER,start=2):
        ws.cell(row=row, column=1, value=users["name"])
        ws.cell(row=row, column=2, value=users["age"])
        ws.cell(row=row, column=3, value=users["email"])
    wb.save(filename)
    file_path = os.path.abspath(filename)
    print("File saved at:", file_path)
    # os.startfile(file_path)
USER=[]
while True:
    print("1.Add user")
    print("2.View user")
    print("3.Analyze data")
    print("4.search user by its name")
    print("5.delete user")
    print("6.Exit")
    Task=input("Enter your choice: ")
    if Task=="1":
        name=input("Enter your name: ")
        try:
            age=int(input("Enter your age: "))
        except:
            print("Please enter a valid age")
            continue
        email=input("Enter your email: ")
        if not is_valid_email(email):
            print("Please enter a valid email")
            continue
        USER.append({"name":name,"age":age,"email":email})
        choice=input("Would you like to add another user? (y/n): ")
        if choice=="n":
            break
    elif Task == "2":

        # with open("USER.json","w")as f:
        #     json.dump(USER,f)
        # with open("USER.json","r")as f:
        #     data=json.load(f)
        # print(data)
        save_users_to_excel(USER)
    elif Task == "3":
        ages=[]
        for a in USER:
            ages.append(a["age"])
            if len(ages) == 0:
                print("No users to analyze")
            else:
                avg_age=sum(ages)/len(ages)
                oldest= max(ages)
                print("average age:",avg_age)
                print("oldest age:",oldest)
    elif Task == "4":
        found = False
        search=input("Enter your search name: ")
        for user in USER:
            if search.lower() == user["name"].lower():
                print("User found:", user)
                found = True
                break
        if not found:
            print("User not found")
    elif Task == "5":
        de_name=input("Enter your user name to be deleted: ")
        de_age = int(input("Enter your user age to be deleted: "))
        de_email = input("Enter your user email to be deleted: ")
        delete_user(de_name,de_age,de_email)

    elif Task == "6":
        break
    else:
        print("Please enter a valid Task")

